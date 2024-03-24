import openpyxl

wb = openpyxl.load_workbook('example.xlsx')

# get all sheet names
# print(wb.sheetnames)

balance_worksheet = wb['Balance']
# print(balance_worksheet)

# read any romdom data like charli slary use ws['alphacolumn laterand rownumber']
# eg: slary column is B and charli row is 4 --> B4
# than use. value to get data
data1 = balance_worksheet['B4'].value
# print(data1)


# get any cell element
nm = balance_worksheet.cell(row=4, column=1).value
bl = balance_worksheet.cell(row=4, column=2).value

# print(f'name {nm} holds {bl} salary')

# get with range many data

values_with_range = balance_worksheet['A2':'B7'] # it returns nested tuple

for a, b in values_with_range:
    print(a.value, b.value)

# close workbook
wb.close()