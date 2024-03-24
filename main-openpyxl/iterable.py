import openpyxl

wb = openpyxl.load_workbook('example.xlsx')

# get all sheet names
# print(wb.sheetnames)

balance_worksheet = wb['Balance']

# iter_rows : it returns tupe for each rows

# rows = balance_worksheet.iter_rows(min_row=1,max_row=7, min_col=1, max_col=2) # it returns generator objects
# # print(rows)

# for row in rows:
#     print(row)


# for a, b in rows:
#     print(a.value, b.value)



# iter_cols : it returns tuple for each column

# columns = balance_worksheet.iter_cols(min_row=1,max_row=7, min_col=1, max_col=2) # it returns generator objects
# print(columns)

# for col in columns:
#     print(col)


# getting all rows

all_rows = list(balance_worksheet.rows)
print(all_rows)


# getting all column

all_columns = list(balance_worksheet.columns)
print(all_columns)