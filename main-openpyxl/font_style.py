import openpyxl
from openpyxl.styles import Color, Font

wb = openpyxl.load_workbook('example.xlsx')

balance_worksheet = wb['Balance']

# for particular element
# font_style = Font(name='Chandas', size='12', color='16aee0', italic=True, bold=True)
# a4 = balance_worksheet['A4']
# a4.font = font_style

rows_count = balance_worksheet.max_row


# for column
font_style = Font(name='Noto Mono', size='13', color='d009eb', italic=True, bold=True)
for i in range(2,rows_count+1):
    balance_worksheet.cell(row=i, column=3).font = font_style

wb.save('example.xlsx')