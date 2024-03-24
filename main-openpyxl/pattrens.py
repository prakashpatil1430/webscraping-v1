import openpyxl
from openpyxl.styles import PatternFill

wb = openpyxl.load_workbook('example.xlsx')

balance_worksheet = wb['Balance']


pattern_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="lightHorizontal")
balance_worksheet["A1"].fill = pattern_fill

wb.save("example.xlsx")


# solid: A solid fill with a single color.
# darkVertical, darkHorizontal, darkDown, darkUp, darkGrid, darkTrellis, darkDiagonalDown, darkDiagonalUp: Various dark patterns.
# lightVertical, lightHorizontal, lightDown, lightUp, lightGrid, lightTrellis, lightDiagonalDown, lightDiagonalUp: Various light patterns.
# mediumGray, darkGray, lightGray: Shades of gray.
# thinHorizontal, thinVertical: Thin horizontal or vertical lines.
# mediumHorizontal, mediumVertical: Medium thickness horizontal or vertical lines.
# thickHorizontal, thickVertical: Thick horizontal or vertical lines.
# thinDiagonalCrosshatch, mediumDiagonalCrosshatch, thickDiagonalCrosshatch, thinDiagonalStripe, thinReverseDiagonalStripe, thinHorizontalStripe, thinVerticalStripe, mediumDiagonalStripe, mediumHorizontalStripe, mediumVerticalStripe, thickDiagonalStripe, thickHorizontalStripe, thickVerticalStripe: Various diagonal and stripe patterns.

