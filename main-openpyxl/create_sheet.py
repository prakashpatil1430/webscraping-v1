from openpyxl import Workbook

# Create a new Workbook object
workbook = Workbook()

"""
    Create sheets named "Balance" and "Score" by default it will save as last sheet
    if you provide index it will save according to requirements
"""
balance_sheet = workbook.create_sheet("Balance")
score_sheet = workbook.create_sheet("Score")


# balance_sheet = workbook.create_sheet("abc", 0)
# score_sheet = workbook.create_sheet("xyz", 2)


# Add data to the "Balance" sheet
balance_sheet.append(["Name", "Salary"])  # Header row
balance_data = [
    ("Alice", 50000),
    ("Bob", 60000),
    ("Charlie", 70000),
    ('pop', 65000),
    ('raj', 25000),
    ('rani', 2000),
    # Add more data as needed
]
for row in balance_data:
    balance_sheet.append(row)

# Add data to the "Score" sheet
score_sheet.append(["Name", "Score"])  # Header row
score_data = [
    ("Alice", 85),
    ("Bob", 90),
    ("Charlie", 75),
    ("pop", 55),
    ("raj", 60),
    ("rani", 75),
    # Add more data as needed
]
for row in score_data:
    score_sheet.append(row)

# Save the workbook to a file
workbook.save(filename='example.xlsx')

# Close the workbook when done
workbook.close()
