import openpyxl

wb = openpyxl.load_workbook('example.xlsx')

balance_worksheet = wb['Balance']

# changing charali salary to 9000
balance_worksheet['B4'] = 900

# to add new data
balance_worksheet['A8'] = 'poorna'
balance_worksheet['B8'] = 33000

# print(balance_worksheet.cell(row=8, column=1).value)
# print(balance_worksheet.cell(row=8, column=2).value)

# another way to update
balance_worksheet.cell(row=8, column=1).value = 'poorna raj'
balance_worksheet.cell(row=8, column=2).value = 65000

# print(balance_worksheet.cell(row=8, column=1).value)
# print(balance_worksheet.cell(row=8, column=2).value)

# create a column doubled amount which adds double of salary

# creating header of third column
balance_worksheet['C1'] = 'Doubled Amount'

rows = balance_worksheet.max_row

for i in range(2, rows + 1):  # Start from row 2 as row 1 is header
    balance_worksheet.cell(row=i, column=3).value = \
          balance_worksheet.cell(row=i, column=2).value * 2

wb.save('example.xlsx')
